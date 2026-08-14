"""Create compact, column-only views of existing SEM coverage tables.

This module never imports or calls the image-analysis pipeline.  It recognizes
the compact schemas emitted by :mod:`export_output_summaries`, selects existing
columns, and writes sibling CSV/XLSX files without recalculating values.
"""

from __future__ import annotations

import argparse
import csv
from copy import copy
from dataclasses import dataclass, field
import os
from pathlib import Path
import shutil
import tempfile
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


COVERAGE_METRICS = ("proj", "caps", "surfw")
REGIONS = ("cap", "homo")
DEFAULT_SUFFIX = "-extracted"

IDENTITY_COLUMNS = ("sample", "source_path", "coverage_branch")

GLOBAL_METADATA = (
    *IDENTITY_COLUMNS,
    "input_image_count",
    "successful_image_count",
    "failed_image_count",
    "detected_bead_count",
    "included_bead_count",
    "sphere_geometry",
)
IMAGE_METADATA = (
    *IDENTITY_COLUMNS,
    "detected_bead_count",
    "included_bead_count",
    "sphere_geometry",
)
ROI_COMMON_METADATA = (
    *IDENTITY_COLUMNS,
    "roi_index",
    "included",
    "exclusion_reasons",
    "sphere_geometry",
)
ROI_REGION_METADATA = {
    "cap": ("cap_valid", "cap_completeness"),
    "homo": ("homo_valid", "local_heterogeneity_valid", "homo_completeness"),
}
LOCAL_CELL_METADATA = (
    *IDENTITY_COLUMNS,
    "roi_index",
    "rotation_offset_deg",
    "radial_band_index",
    "polar_sector_index",
    "radial_inner_fraction",
    "radial_outer_fraction",
    "polar_start_deg",
    "polar_end_deg",
    "valid",
    "completeness",
    "reference_pixel_count",
    "ag_pixel_count",
)
RADIAL_PROFILE_METADATA = (
    *IDENTITY_COLUMNS,
    "roi_index",
    "radial_band_index",
    "radial_inner_fraction",
    "radial_outer_fraction",
    "valid",
    "completeness",
)
POLAR_SECTOR_METADATA = (
    *IDENTITY_COLUMNS,
    "roi_index",
    "rotation_index",
    "rotation_offset_deg",
    "polar_sector_index",
    "polar_start_deg",
    "polar_end_deg",
    "valid",
    "completeness",
)

_HETEROGENEITY_FAMILIES = ("rad", "polar", "local_total", "local_residual")


def _global_columns() -> tuple[str, ...]:
    metric_columns: list[str] = []
    for region in REGIONS:
        for metric in COVERAGE_METRICS:
            metric_columns.extend(
                (f"{region}_{metric}_mean_pct", f"{region}_{metric}_median_pct", f"{region}_{metric}_sd_pp")
            )
    for family in _HETEROGENEITY_FAMILIES:
        for metric in COVERAGE_METRICS:
            metric_columns.extend(
                (
                    f"{family}_{metric}_mean_pp",
                    f"{family}_{metric}_median_pp",
                    f"{family}_{metric}_between_bead_sd_pp",
                )
            )
    metric_columns.extend(f"cap_sens_{metric}_median_pp" for metric in COVERAGE_METRICS)
    return (
        *GLOBAL_METADATA,
        "selected_sphere_diameter_mean_um",
        "selected_sphere_diameter_sd_um",
        *metric_columns,
    )


def _image_columns() -> tuple[str, ...]:
    metric_columns: list[str] = []
    for region in REGIONS:
        metric_columns.extend(f"{region}_{metric}_mean_pct" for metric in COVERAGE_METRICS)
    for family in _HETEROGENEITY_FAMILIES:
        metric_columns.extend(f"{family}_{metric}_mean_pp" for metric in COVERAGE_METRICS)
    metric_columns.extend(f"cap_sens_{metric}_mean_pp" for metric in COVERAGE_METRICS)
    return (*IMAGE_METADATA, "selected_sphere_diameter_mean_um", *metric_columns)


def _roi_columns() -> tuple[str, ...]:
    metric_columns: list[str] = []
    for region in REGIONS:
        metric_columns.extend(f"{region}_{metric}_pct" for metric in COVERAGE_METRICS)
    for family in _HETEROGENEITY_FAMILIES:
        metric_columns.extend(f"{family}_{metric}_pp" for metric in COVERAGE_METRICS)
    metric_columns.extend(f"rad_slope_{metric}_pp_per_R" for metric in COVERAGE_METRICS)
    metric_columns.extend(f"cap_sens_{metric}_pp" for metric in COVERAGE_METRICS)
    return (
        *ROI_COMMON_METADATA[:6],
        "cap_valid",
        "homo_valid",
        "local_heterogeneity_valid",
        "cap_completeness",
        "homo_completeness",
        "sphere_geometry",
        "diam_xy_mean_um",
        "diam_eq_um",
        "diam_inscribed_um",
        "selected_sphere_diameter_um",
        *metric_columns,
    )


def _global_descriptive_population_medians() -> frozenset[str]:
    """Columns that duplicate a mean over the same exported bead population.

    The polar/total/residual inputs summarized by these columns are themselves
    robust per-bead estimators (medians across grid rotations).  The compact
    global ``*_median_pp`` columns add a second, descriptive median *across
    beads* beside ``*_mean_pp`` and are therefore optional.  ROI-level robust
    values and the global/image means of those values are not in this set.
    """

    return frozenset(
        {
            *(
                f"{region}_{metric}_median_pct"
                for region in REGIONS
                for metric in COVERAGE_METRICS
            ),
            *(
                f"{family}_{metric}_median_pp"
                for family in _HETEROGENEITY_FAMILIES
                for metric in COVERAGE_METRICS
            ),
        }
    )


@dataclass(frozen=True)
class TableSchema:
    key: str
    columns: tuple[str, ...]
    fingerprint: frozenset[str]
    filenames: frozenset[str]
    sheet_names: frozenset[str]
    regions: frozenset[str]
    descriptive_population_medians: frozenset[str] = frozenset()
    column_selectable: bool = True


SCHEMAS = (
    TableSchema(
        "coverage_samples",
        _global_columns(),
        frozenset({"sample", "source_path", "input_image_count", "cap_proj_mean_pct", "homo_proj_mean_pct"}),
        frozenset({"coverage_global_summaries.csv"}),
        frozenset({"Coverage", "Coverage samples"}),
        frozenset(REGIONS),
        descriptive_population_medians=_global_descriptive_population_medians(),
    ),
    TableSchema(
        "coverage_images",
        _image_columns(),
        frozenset({"sample", "source_path", "detected_bead_count", "cap_proj_mean_pct", "homo_proj_mean_pct"}),
        frozenset({"coverage_image_summaries.csv"}),
        frozenset({"Coverage images"}),
        frozenset(REGIONS),
    ),
    TableSchema(
        "coverage_rois",
        _roi_columns(),
        frozenset({"sample", "source_path", "roi_index", "included", "cap_proj_pct", "homo_proj_pct"}),
        frozenset({"coverage_roi_details.csv"}),
        frozenset({"Coverage ROIs"}),
        frozenset(REGIONS),
    ),
    TableSchema(
        "local_cells",
        (*LOCAL_CELL_METADATA, *(f"{metric}_pct" for metric in COVERAGE_METRICS)),
        frozenset({"sample", "source_path", "roi_index", "radial_band_index", "polar_sector_index", "reference_pixel_count", "proj_pct"}),
        frozenset({"coverage_local_cell_details.csv"}),
        frozenset({"Local cells"}),
        frozenset({"homo"}),
    ),
    TableSchema(
        "radial_profile",
        (
            *IDENTITY_COLUMNS,
            "roi_index",
            "radial_band_index",
            "radial_inner_fraction",
            "radial_outer_fraction",
            *(f"radial_center_{metric}_fraction" for metric in COVERAGE_METRICS),
            "valid",
            "completeness",
            *(f"{metric}_pct" for metric in COVERAGE_METRICS),
        ),
        frozenset({"sample", "source_path", "roi_index", "radial_band_index", "radial_center_proj_fraction", "proj_pct"}),
        frozenset({"coverage_radial_profile_details.csv"}),
        frozenset({"Radial profile"}),
        frozenset({"homo"}),
    ),
    TableSchema(
        "polar_sectors",
        (*POLAR_SECTOR_METADATA, *(f"{metric}_pct" for metric in COVERAGE_METRICS)),
        frozenset({"sample", "source_path", "roi_index", "rotation_index", "polar_sector_index", "proj_pct"}),
        frozenset({"coverage_polar_sector_details.csv"}),
        frozenset({"Polar sectors"}),
        frozenset({"homo"}),
    ),
    TableSchema(
        "polar_rotations",
        (
            *IDENTITY_COLUMNS,
            "roi_index",
            "metric",
            "rotation_index",
            "rotation_offset_deg",
            "valid_sector_count",
            "valid_cell_count",
            "polar_sd_pp",
            "polar_mad_pp",
            "local_total_sd_pp",
            "local_total_mad_pp",
            "local_residual_sd_pp",
            "local_residual_mad_pp",
            "reconstructed_coverage_pct",
            "reconstruction_delta_pp",
        ),
        frozenset({"sample", "source_path", "roi_index", "metric", "rotation_index", "polar_sd_pp"}),
        frozenset({"coverage_polar_rotation_summaries.csv"}),
        frozenset({"Polar rotations"}),
        frozenset({"homo"}),
        column_selectable=False,
    ),
)

KNOWN_UNSELECTABLE_FILENAMES = frozenset({"coverage_failures.csv"})
KNOWN_UNSELECTABLE_SHEETS = frozenset({"Coverage failures"})
LEGACY_COVERAGE_FINGERPRINT = frozenset(
    {"name", "json_file", "mean_coverage", "mean_coverage_percent", "included_roi_count"}
)


@dataclass(frozen=True)
class Selection:
    schema: TableSchema
    retained_columns: tuple[str, ...]


@dataclass
class RunSummary:
    files_scanned: int = 0
    recognized_tables: int = 0
    files_written: int = 0
    unrelated_skipped: int = 0
    existing_outputs_skipped: int = 0
    self_outputs_skipped: int = 0
    errors: int = 0
    messages: list[str] = field(default_factory=list)


def validate_suffix(suffix: str) -> str:
    if not suffix:
        raise argparse.ArgumentTypeError("suffix must not be empty")
    if suffix in {".", ".."} or "/" in suffix or "\\" in suffix:
        raise argparse.ArgumentTypeError("suffix must be a filename-stem suffix, not a path")
    return suffix


def build_output_path(source: Path, suffix: str = DEFAULT_SUFFIX) -> Path:
    return source.with_name(f"{source.stem}{suffix}{source.suffix}")


def is_temporary_table(path: Path) -> bool:
    name = path.name
    return (
        name.startswith("~$")
        or name.startswith(".")
        or path.stem.endswith((".tmp", ".part", ".bak"))
    )


def discover_supported_files(root: Path, suffix: str = DEFAULT_SUFFIX) -> list[Path]:
    """Return recursive CSV/XLSX inputs in deterministic relative-path order."""

    paths: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.casefold() not in {".csv", ".xlsx"}:
            continue
        if is_temporary_table(path) or path.stem.endswith(suffix):
            continue
        paths.append(path)
    return sorted(paths, key=lambda item: item.relative_to(root).as_posix().casefold())


def classify_table(headers: Sequence[object], *, filename: str = "", sheet_name: str = "") -> TableSchema | None:
    """Match one complete current export schema using explicit fingerprints."""

    if not headers or any(not isinstance(value, str) or not value for value in headers):
        return None
    if len(set(headers)) != len(headers):
        return None
    header_set = set(headers)
    matches = [
        schema
        for schema in SCHEMAS
        if schema.fingerprint <= header_set and header_set <= set(schema.columns)
    ]
    if len(matches) == 1:
        return matches[0]
    # Known names do not override an unsafe or ambiguous header match.
    named = [
        schema
        for schema in SCHEMAS
        if filename in schema.filenames or sheet_name in schema.sheet_names
    ]
    if len(named) == 1 and named[0].fingerprint <= header_set and header_set <= set(named[0].columns):
        return named[0]
    return None


def _selected_scientific_columns(schema: TableSchema, metric: str, region: str) -> frozenset[str]:
    if schema.key == "coverage_samples":
        if region == "cap":
            return frozenset(
                {
                    f"cap_{metric}_mean_pct",
                    f"cap_{metric}_median_pct",
                    f"cap_{metric}_sd_pp",
                    f"cap_sens_{metric}_median_pp",
                }
            )
        return frozenset(
            {
                f"homo_{metric}_mean_pct",
                f"homo_{metric}_median_pct",
                f"homo_{metric}_sd_pp",
                *(
                    f"{family}_{metric}_{stat}"
                    for family in _HETEROGENEITY_FAMILIES
                    for stat in ("mean_pp", "median_pp", "between_bead_sd_pp")
                ),
            }
        )
    if schema.key == "coverage_images":
        if region == "cap":
            return frozenset({f"cap_{metric}_mean_pct", f"cap_sens_{metric}_mean_pp"})
        return frozenset(
            {f"homo_{metric}_mean_pct", *(f"{family}_{metric}_mean_pp" for family in _HETEROGENEITY_FAMILIES)}
        )
    if schema.key == "coverage_rois":
        if region == "cap":
            return frozenset({f"cap_{metric}_pct", f"cap_sens_{metric}_pp"})
        return frozenset(
            {
                f"homo_{metric}_pct",
                *(f"{family}_{metric}_pp" for family in _HETEROGENEITY_FAMILIES),
                f"rad_slope_{metric}_pp_per_R",
            }
        )
    if schema.key == "local_cells":
        return frozenset({f"{metric}_pct"})
    if schema.key == "radial_profile":
        return frozenset({f"radial_center_{metric}_fraction", f"{metric}_pct"})
    if schema.key == "polar_sectors":
        return frozenset({f"{metric}_pct"})
    return frozenset()


def _metadata_columns(schema: TableSchema, region: str) -> frozenset[str]:
    if schema.key == "coverage_samples":
        return frozenset(GLOBAL_METADATA)
    if schema.key == "coverage_images":
        return frozenset(IMAGE_METADATA)
    if schema.key == "coverage_rois":
        return frozenset((*ROI_COMMON_METADATA, *ROI_REGION_METADATA[region]))
    if schema.key == "local_cells":
        return frozenset(LOCAL_CELL_METADATA)
    if schema.key == "radial_profile":
        return frozenset(RADIAL_PROFILE_METADATA)
    if schema.key == "polar_sectors":
        return frozenset(POLAR_SECTOR_METADATA)
    return frozenset()


def select_columns(
    headers: Sequence[str],
    schema: TableSchema,
    *,
    metric: str = "caps",
    region: str = "homo",
    drop_medians: bool = False,
) -> Selection | None:
    """Return an ordered column-only selection, or ``None`` when inapplicable."""

    if metric not in COVERAGE_METRICS:
        raise ValueError(f"Unsupported coverage metric: {metric}")
    if region not in REGIONS:
        raise ValueError(f"Unsupported region/domain: {region}")
    if region not in schema.regions or not schema.column_selectable:
        return None
    scientific = _selected_scientific_columns(schema, metric, region)
    if not scientific.intersection(headers):
        return None
    keep = _metadata_columns(schema, region) | scientific
    if drop_medians:
        keep -= schema.descriptive_population_medians
    retained = tuple(column for column in headers if column in keep)
    return Selection(schema=schema, retained_columns=retained)


def _csv_headers(path: Path) -> tuple[list[str], str]:
    with path.open("rb") as raw_handle:
        has_bom = raw_handle.read(3) == b"\xef\xbb\xbf"
    encoding = "utf-8-sig" if has_bom else "utf-8"
    with path.open("r", newline="", encoding=encoding) as handle:
        return next(csv.reader(handle), []), encoding


def _write_selected_csv(source: Path, output: Path, retained: Sequence[str], encoding: str) -> None:
    with source.open("r", newline="", encoding=encoding) as source_handle:
        reader = csv.reader(source_handle)
        headers = next(reader)
        indices = [headers.index(column) for column in retained]
        with tempfile.NamedTemporaryFile(
            "w", newline="", encoding=encoding, dir=output.parent, suffix=".csv", delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
            writer = csv.writer(temporary)
            writer.writerow(retained)
            for row in reader:
                writer.writerow([row[index] if index < len(row) else "" for index in indices])
    try:
        os.replace(temporary_path, output)
    except BaseException:
        temporary_path.unlink(missing_ok=True)
        raise


def _sheet_headers(worksheet) -> list[str]:
    values = [cell.value for cell in worksheet[1]]
    while values and values[-1] is None:
        values.pop()
    return values


def _delete_unretained_columns(worksheet, headers: Sequence[str], retained: Sequence[str]) -> None:
    keep = set(retained)
    widths = {
        header: copy(worksheet.column_dimensions[get_column_letter(index)])
        for index, header in enumerate(headers, start=1)
    }
    for index in range(len(headers), 0, -1):
        if headers[index - 1] not in keep:
            worksheet.delete_cols(index, 1)
    for key in list(worksheet.column_dimensions):
        del worksheet.column_dimensions[key]
    for index, header in enumerate(retained, start=1):
        dimension = widths[header]
        dimension.index = get_column_letter(index)
        dimension.min = index
        dimension.max = index
        worksheet.column_dimensions[get_column_letter(index)] = dimension
    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(retained))}{worksheet.max_row}"
    for table in worksheet.tables.values():
        table.ref = f"A1:{get_column_letter(len(retained))}{worksheet.max_row}"


def _write_selected_workbook(
    source: Path,
    output: Path,
    selections: dict[str, Selection],
    omitted_sheets: set[str],
) -> None:
    with tempfile.NamedTemporaryFile(dir=output.parent, suffix=".xlsx", delete=False) as temporary:
        temporary_path = Path(temporary.name)
    try:
        shutil.copy2(source, temporary_path)
        workbook = load_workbook(temporary_path, data_only=False)
        for sheet_name in list(workbook.sheetnames):
            if sheet_name in omitted_sheets:
                del workbook[sheet_name]
                continue
            selection = selections.get(sheet_name)
            if selection is None:
                continue
            worksheet = workbook[sheet_name]
            _delete_unretained_columns(
                worksheet,
                _sheet_headers(worksheet),
                selection.retained_columns,
            )
        workbook.save(temporary_path)
        workbook.close()
        os.replace(temporary_path, output)
    except BaseException:
        temporary_path.unlink(missing_ok=True)
        raise


def _is_known_unselectable_csv(path: Path, headers: Sequence[str]) -> bool:
    return path.name in KNOWN_UNSELECTABLE_FILENAMES or LEGACY_COVERAGE_FINGERPRINT <= set(headers)


def _process_csv(
    path: Path,
    *,
    metric: str,
    region: str,
    suffix: str,
    dry_run: bool,
    overwrite: bool,
    drop_medians: bool,
    summary: RunSummary,
) -> None:
    headers, encoding = _csv_headers(path)
    schema = classify_table(headers, filename=path.name)
    if schema is None:
        summary.unrelated_skipped += 1
        if _is_known_unselectable_csv(path, headers):
            summary.messages.append(f"SKIP {path}: recognized coverage table has no selectable {metric}+{region} columns")
        else:
            summary.messages.append(f"SKIP {path}: unrelated or unrecognized table schema")
        return
    summary.recognized_tables += 1
    selection = select_columns(
        headers,
        schema,
        metric=metric,
        region=region,
        drop_medians=drop_medians,
    )
    if selection is None:
        summary.unrelated_skipped += 1
        summary.messages.append(
            f"SKIP {path}: {schema.key} is not safely column-selectable for metric={metric}, region={region}"
        )
        return
    output = build_output_path(path, suffix)
    summary.messages.append(
        f"{'DRY-RUN' if dry_run else 'PROCESS'} {path}: {len(headers)} -> {len(selection.retained_columns)} columns; metric={metric}; region={region}; drop-medians={'yes' if drop_medians else 'no'}"
    )
    if dry_run:
        return
    if output.exists() and not overwrite:
        summary.existing_outputs_skipped += 1
        summary.messages.append(f"SKIP {output}: output exists (use --overwrite)")
        return
    _write_selected_csv(path, output, selection.retained_columns, encoding)
    summary.files_written += 1


def _process_xlsx(
    path: Path,
    *,
    metric: str,
    region: str,
    suffix: str,
    dry_run: bool,
    overwrite: bool,
    drop_medians: bool,
    summary: RunSummary,
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    selections: dict[str, Selection] = {}
    omitted: set[str] = set()
    details: list[str] = []
    recognized_here = 0
    try:
        for worksheet in workbook.worksheets:
            headers = _sheet_headers(worksheet)
            schema = classify_table(headers, filename=path.name, sheet_name=worksheet.title)
            if schema is None:
                if worksheet.title in KNOWN_UNSELECTABLE_SHEETS:
                    omitted.add(worksheet.title)
                    details.append(f"  omit {worksheet.title}: no selectable scientific metric/domain columns")
                continue
            recognized_here += 1
            selection = select_columns(
                headers,
                schema,
                metric=metric,
                region=region,
                drop_medians=drop_medians,
            )
            if selection is None:
                omitted.add(worksheet.title)
                details.append(
                    f"  omit {worksheet.title}: {schema.key} is not safely column-selectable for metric={metric}, region={region}"
                )
                continue
            selections[worksheet.title] = selection
            details.append(f"  {worksheet.title}: {len(headers)} -> {len(selection.retained_columns)} columns")
    finally:
        workbook.close()
    summary.recognized_tables += recognized_here
    if not selections:
        summary.unrelated_skipped += 1
        reason = "recognized sheets are not applicable" if recognized_here or omitted else "unrelated or unrecognized workbook"
        summary.messages.append(f"SKIP {path}: {reason}")
        summary.messages.extend(details)
        return
    output = build_output_path(path, suffix)
    summary.messages.append(
        f"{'DRY-RUN' if dry_run else 'PROCESS'} {path}: metric={metric}; region={region}; drop-medians={'yes' if drop_medians else 'no'}"
    )
    summary.messages.extend(details)
    if dry_run:
        return
    if output.exists() and not overwrite:
        summary.existing_outputs_skipped += 1
        summary.messages.append(f"SKIP {output}: output exists (use --overwrite)")
        return
    _write_selected_workbook(path, output, selections, omitted)
    summary.files_written += 1


def extract_tables(
    root: Path,
    *,
    coverage_metric: str = "caps",
    region: str = "homo",
    suffix: str = DEFAULT_SUFFIX,
    dry_run: bool = False,
    overwrite: bool = False,
    drop_medians: bool = False,
) -> RunSummary:
    """Recursively extract recognized tables and return a testable run summary."""

    root = root.expanduser().resolve()
    if not root.is_dir():
        raise ValueError(f"Root directory does not exist or is not a directory: {root}")
    validate_suffix(suffix)
    if coverage_metric not in COVERAGE_METRICS:
        raise ValueError(f"Unsupported coverage metric: {coverage_metric}")
    if region not in REGIONS:
        raise ValueError(f"Unsupported region/domain: {region}")
    summary = RunSummary()
    all_candidates = [
        path
        for path in root.rglob("*")
        if path.is_file() and path.suffix.casefold() in {".csv", ".xlsx"}
    ]
    paths = discover_supported_files(root, suffix)
    summary.self_outputs_skipped = sum(
        1
        for path in all_candidates
        if not is_temporary_table(path) and path.stem.endswith(suffix)
    )
    summary.files_scanned = len(paths)
    for path in paths:
        try:
            if path.suffix.casefold() == ".csv":
                _process_csv(
                    path,
                    metric=coverage_metric,
                    region=region,
                    suffix=suffix,
                    dry_run=dry_run,
                    overwrite=overwrite,
                    drop_medians=drop_medians,
                    summary=summary,
                )
            else:
                _process_xlsx(
                    path,
                    metric=coverage_metric,
                    region=region,
                    suffix=suffix,
                    dry_run=dry_run,
                    overwrite=overwrite,
                    drop_medians=drop_medians,
                    summary=summary,
                )
        except Exception as exc:  # Continue safely through large result trees.
            summary.errors += 1
            summary.messages.append(f"ERROR {path}: {exc}")
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Recursively create compact sibling views of existing SEM coverage CSV/XLSX tables. "
            "No scientific values are recalculated."
        )
    )
    parser.add_argument("root", type=Path, help="Root directory recursively containing existing result tables.")
    parser.add_argument(
        "--coverage-metric",
        choices=COVERAGE_METRICS,
        default="caps",
        help="Existing coverage metric family to retain (default: %(default)s).",
    )
    parser.add_argument(
        "--region",
        choices=REGIONS,
        default="homo",
        help="Existing bead region/domain to retain: cap or homogeneity-domain annulus (default: %(default)s).",
    )
    parser.add_argument(
        "--suffix",
        type=validate_suffix,
        default=DEFAULT_SUFFIX,
        help="Sibling filename-stem suffix (default: %(default)s).",
    )
    parser.add_argument("--dry-run", action="store_true", help="Inspect and report without writing files.")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing extracted siblings.")
    parser.add_argument(
        "--drop-medians",
        action="store_true",
        help=(
            "Drop only schema-classified descriptive population medians that duplicate a mean; "
            "retain rotation-robust scientific estimators and medians without an equivalent mean."
        ),
    )
    return parser


def _print_summary(summary: RunSummary) -> None:
    for message in summary.messages:
        print(message)
    print("Summary:")
    print(f"  files scanned: {summary.files_scanned}")
    print(f"  recognized tables: {summary.recognized_tables}")
    print(f"  files written: {summary.files_written}")
    print(f"  files skipped as unrelated/inapplicable: {summary.unrelated_skipped}")
    print(f"  existing outputs skipped: {summary.existing_outputs_skipped}")
    print(f"  prior extracted files ignored: {summary.self_outputs_skipped}")
    print(f"  errors: {summary.errors}")


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        summary = extract_tables(
            args.root,
            coverage_metric=args.coverage_metric,
            region=args.region,
            suffix=args.suffix,
            dry_run=args.dry_run,
            overwrite=args.overwrite,
            drop_medians=args.drop_medians,
        )
    except ValueError as exc:
        parser.error(str(exc))
    _print_summary(summary)
    return 1 if summary.errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
